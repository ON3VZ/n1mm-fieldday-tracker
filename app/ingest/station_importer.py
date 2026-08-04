"""Import the participant list from Excel (.xlsx) or CSV (§7).

Column recognition is header-based, case- and whitespace-insensitive, with
synonyms (§7.1). Band columns are any header that parses as a valid band via
``band_plan.parse_band_label``; they determine the *proposal* for the field
day's ``selected_bands``, but their cell contents are never imported — the
matrix always starts empty (BR-10).

Validation (§7.2) never blocks the import: problems are reported per row in
the import report. Duplicates after normalization are reported explicitly.
"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from pathlib import Path

from app.core.band_plan import parse_band_label
from app.core.callsign import normalize_callsign
from app.core.models import Station, StationSource

logger = logging.getLogger(__name__)

# Normalized header → model field. Header normalization: lowercase, trimmed,
# trailing dots removed, inner whitespace collapsed.
_HEADER_SYNONYMS: dict[str, str] = {
    "call": "callsign",
    "callsign": "callsign",
    "roepnaam": "callsign",
    "categorie": "category",
    "category": "category",
    "sectie": "section",
    "section": "section",
    "opm": "remarks",
    "opmerking": "remarks",
    "opmerkingen": "remarks",
    "remarks": "remarks",
    "name": "name",
    "naam": "name",
    "club": "club",
}


# §7.1 / fase 26: the participant list has a FIXED set of mandatory columns.
# Order is free and synonyms are accepted, but these must all be present,
# otherwise the file is refused as a whole with an explicit layout hint.
REQUIRED_FIELDS: tuple[str, ...] = ("callsign", "category", "section")

# Human-readable header names per required field, for the error message.
_REQUIRED_LABELS: dict[str, str] = {
    "callsign": "Call",
    "category": "categorie",
    "section": "sectie",
}


def expected_format() -> dict:
    """The layout the importer expects, for display in the UI (fase 26).

    Returned as data, not as a sentence, so the view can render it as a
    table in the operator's own language without duplicating the rules.
    """
    return {
        "required": [
            {"header": "Call", "synonyms": ["Callsign", "Roepnaam"],
             "example": "ON4BAF/P"},
            {"header": "categorie", "synonyms": ["Category"],
             "example": "Open All Band Low Power"},
            {"header": "sectie", "synonyms": ["Section"], "example": "RST"},
        ],
        "required_bands": {
            "header": "40M / 80M / 160M ...",
            "example": "",
        },
        "optional": [
            {"header": "Nummer", "example": "1"},
            {"header": "Naam", "synonyms": ["Name"], "example": ""},
            {"header": "Club", "example": ""},
            {"header": "Opm.", "synonyms": ["Remarks", "Opmerking"], "example": ""},
        ],
        "notes": [
            "header_first_row",
            "order_free",
            "band_cells_ignored",
        ],
    }


@dataclass
class RowIssue:
    """One reported problem for one input row."""

    row_number: int  # 1-based, as shown in Excel
    callsign: str
    reason: str


@dataclass
class StationImportResult:
    """Report of one import run (§7.2). Errors never block the import."""

    stations: list[Station] = field(default_factory=list)
    band_columns: list[str] = field(default_factory=list)  # proposal for selected_bands
    issues: list[RowIssue] = field(default_factory=list)
    rows_read: int = 0
    source: str = ""
    # Fase 26: set when the file does not match the fixed format. Nothing is
    # imported in that case; the caller shows the expected layout instead.
    missing_columns: list[str] = field(default_factory=list)
    found_headers: list[str] = field(default_factory=list)

    @property
    def format_ok(self) -> bool:
        return not self.missing_columns

    def to_report_dict(self) -> dict:
        """Summary for the sync log / UI."""
        return {
            "source": self.source,
            "rows_read": self.rows_read,
            "imported": len(self.stations),
            "band_columns": list(self.band_columns),
            "missing_columns": list(self.missing_columns),
            "found_headers": list(self.found_headers),
            "issues": [
                {"row": i.row_number, "callsign": i.callsign, "reason": i.reason}
                for i in self.issues
            ],
        }


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).split()).strip().lower()
    return text.rstrip(".")


def _map_headers(headers: list) -> tuple[dict[int, str], list[str]]:
    """Map column index → model field, and collect band columns in order.

    Unknown columns (e.g. ``Nummer``) are ignored.
    """
    mapping: dict[int, str] = {}
    band_columns: list[str] = []
    for idx, raw in enumerate(headers):
        norm = _normalize_header(raw)
        if not norm:
            continue
        if norm in _HEADER_SYNONYMS:
            fieldname = _HEADER_SYNONYMS[norm]
            if fieldname not in mapping.values():  # first occurrence wins
                mapping[idx] = fieldname
            continue
        band = parse_band_label(str(raw))
        if band is not None and band not in band_columns:
            band_columns.append(band)
    return mapping, band_columns


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_result(
    rows: list[list],
    source: StationSource,
    strict: bool,
    first_data_row_number: int,
) -> StationImportResult:
    result = StationImportResult(source=source.value)
    if not rows:
        result.issues.append(RowIssue(1, "", "no data found"))
        return result

    header_map, band_columns = _map_headers(rows[0])
    result.band_columns = band_columns
    result.found_headers = [
        str(h).strip() for h in rows[0] if h is not None and str(h).strip()
    ]

    # Fixed-format check (fase 26): all mandatory columns plus at least one
    # band column must be present, otherwise nothing is imported at all.
    present = set(header_map.values())
    missing = [
        _REQUIRED_LABELS[fieldname]
        for fieldname in REQUIRED_FIELDS
        if fieldname not in present
    ]
    if not band_columns:
        missing.append("band (40M / 80M / 160M ...)")
    if missing:
        result.missing_columns = missing
        result.issues.append(
            RowIssue(1, "", "file layout does not match: missing column(s) "
                     + ", ".join(missing))
        )
        result.stations = []
        result.band_columns = []
        return result

    seen: dict[str, tuple[int, str]] = {}  # normalized → (row_number, original)
    for offset, row in enumerate(rows[1:]):
        row_number = first_data_row_number + offset
        values = {fieldname: "" for fieldname in _HEADER_SYNONYMS.values()}
        for idx, fieldname in header_map.items():
            if idx < len(row):
                values[fieldname] = _cell_str(row[idx])

        callsign = values["callsign"]
        if not callsign and not any(values.values()):
            continue  # entirely empty row: skip silently
        result.rows_read += 1

        if not callsign:
            result.issues.append(RowIssue(row_number, "", "callsign is missing"))
            continue

        normalized = normalize_callsign(callsign, strict=strict)
        if normalized is None:
            result.issues.append(
                RowIssue(row_number, callsign, "not a plausible callsign")
            )
            continue

        if normalized in seen:
            prev_row, prev_original = seen[normalized]
            result.issues.append(
                RowIssue(
                    row_number,
                    callsign,
                    f"duplicate after normalization: same station as "
                    f"{prev_original!r} on row {prev_row} (normalized: {normalized})",
                )
            )
            continue
        seen[normalized] = (row_number, callsign)

        result.stations.append(
            Station(
                original_callsign=callsign,
                normalized_callsign=normalized,
                name=values["name"],
                club=values["club"],
                category=values["category"],
                section=values["section"],
                remarks=values["remarks"],
                source=source,
            )
        )
    return result


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def import_stations_from_excel(
    path: Path | str,
    strict: bool = False,
    sheet_name: str | None = None,
) -> StationImportResult:
    """Import stations from an ``.xlsx`` file.

    Uses the given sheet, or the first sheet when *sheet_name* is None.
    Raises ``ValueError`` when the file cannot be opened as Excel; row-level
    problems go into the report instead.
    """
    from openpyxl import load_workbook  # imported lazily: optional at runtime

    path = Path(path)
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file {path.name}: {exc}") from exc

    try:
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet {sheet_name!r} not found in {path.name}; "
                    f"available: {', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    return _rows_to_result(
        rows, StationSource.EXCEL, strict=strict, first_data_row_number=2
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def import_stations_from_csv(path: Path | str, strict: bool = False) -> StationImportResult:
    """Import stations from a CSV file with the same column logic (§7.1).

    Delimiter is auto-detected (``,`` or ``;`` — Belgian Excel exports use
    ``;``); encoding UTF-8 with BOM tolerance.
    """
    path = Path(path)
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        # Legacy exports: fall back to Latin-1, which never fails.
        text = path.read_text(encoding="latin-1")
    except OSError as exc:
        raise ValueError(f"Cannot read CSV file {path.name}: {exc}") from exc

    lines = text.splitlines()
    if not lines:
        result = StationImportResult(source=StationSource.CSV.value)
        result.issues.append(RowIssue(1, "", "no data found"))
        return result

    try:
        dialect = csv.Sniffer().sniff(lines[0], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel  # default comma

    reader = csv.reader(lines, dialect)
    rows = [list(row) for row in reader]
    return _rows_to_result(
        rows, StationSource.CSV, strict=strict, first_data_row_number=2
    )
